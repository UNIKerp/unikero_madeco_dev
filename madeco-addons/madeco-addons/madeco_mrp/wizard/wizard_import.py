# -*- coding: utf-8 -*-

from odoo import models, fields, api
import logging
logger = logging.getLogger(__name__)
import openpyxl
import io
import base64
from datetime import datetime

class wizard_pdp_import(models.TransientModel):
	_name = 'pdp.import.wizard'
	_inherit = 'pdp.import.wizard'
	
	
	
	def importer(self):
		# vérifier les données et le format de fichier
		data = base64.b64decode(self.fichier)
		output = open('/tmp/pdp_import.xlsx','wb')
		output.write(data)
		output.close()
				
		wb = openpyxl.load_workbook(filename='/tmp/pdp_import.xlsx')
		
		if 'PDP' not in wb.sheetnames:
			raise UserError('Aucun onglet du nom de PDP dans le fichier, import impossible')
		
		ws = wb['PDP']
		data = []
		for row in ws.iter_rows(min_row=1):
			line = []
			for cell in row:
				if type(cell.value) is str:
					line.append(cell.value.strip())
				else:
					line.append(cell.value)
			data.append(line)
		
		logger.info(data)
		
		self.env['mrp.production.schedule'].search([]).unlink()
		i = 1
		dates = []
		demandes = []
		reappros = []
		warehouses = []
		for line in data:
			# première ligne = dates répétées 5 fois
			if i == 1:
				nb_dates = int((len(line) - 3 ) / 5)

				for j in range(3,nb_dates*5,5):			
					dates.append(datetime.strptime(line[j],'%d-%m-%Y'))
		
			if i > 2:
				# première colonne = nom produit
				#product_id = self.env['product.product'].search([('id','=',line[0])],limit=1)
				product_id = self.env['product.product'].search([('default_code','=',line[0])],limit=1)
				if len(product_id) == 0:
					raise UserError("Le produit dont le nom est {} n'a pas été trouvé dans Odoo".format(line[0]))
				
				demandes = []
				reappros = []
				warehouses = []
				
				for j in range(3,nb_dates*5,5):
					logger.info(j)
					logger.info(j+3)
					logger.info(line)
					demandes.append(line[1+j])
					reappros.append(line[2+j])
					warehouse = self.env['stock.warehouse'].search([('id','=',line[4+j])])
					if len(warehouse)>0:
						warehouses.append(warehouse.id)
					else:
						warehouses.append(self.env['mrp.production.schedule']._default_warehouse_id().id)
				
				logger.info(demandes)
				logger.info(reappros)
					
				forecast_ids = []
				for j in range(len(dates)):
					value_forecast = {
						'date': dates[j],
						'forecast_qty': demandes[j],
						'replenish_qty': reappros[j],
					}
					forecast_ids.append((0,0,value_forecast))				
				
				value_schedule = {
					'product_id': product_id.id,
					'warehouse_id': warehouses[0],
					'forecast_ids': forecast_ids,
				}
				logger.info(value_schedule)
				self.env['mrp.production.schedule'].create(value_schedule)
			i += 1
		
		return {
			'type': 'ir.actions.client',
			'tag': 'reload',
		}