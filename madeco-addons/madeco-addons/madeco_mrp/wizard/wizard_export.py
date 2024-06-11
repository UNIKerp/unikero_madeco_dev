# -*- coding: utf-8 -*-

from odoo import models, fields, api
import logging
logger = logging.getLogger(__name__)
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import base64
from tempfile import TemporaryFile


class wizard_pdp_export(models.TransientModel):
	_name = 'pdp.export.wizard'
	_inherit = 'pdp.export.wizard'
	
	def exporter(self):
		forecast_obj = self.env['mrp.product.forecast']
		logger.info(self.production_schedule_ids)
		"""
		export_xls = [
			[ 'date1', 'date2', 'date3', ...], <-- lecture seule
			['nom produit', qte1, qte2, qte3 ...], <-- lecture seule
			['demande planifiée', forecast_qty1, forecast_qty2, ...],
			['réappro suggéré', replenish_qty1, replenish_qty2, ...],
			['stock planifié', qte1 - forecast_qty1 + replenish_qty1,...] <-- lecture seule
			['Entrepot ID','warehouse_id']
		]
		"""
		export_xls = []
		header = False
		for production_schedule_id in self.production_schedule_ids:
			logger.info(production_schedule_id)
			
			logger.info(production_schedule_id.get_production_schedule_view_state())
			production_schedule_id_forecast_ids = production_schedule_id.get_production_schedule_view_state()[0]['forecast_ids']
			if not header:
				fournisseur = self.rechercher_fournisseur(production_schedule_id.product_id)
				line = [
					['','',''],
					['','',''],
					['{}'.format(production_schedule_id.product_id.default_code),'{}-{}'.format(production_schedule_id.product_id.default_code,production_schedule_id.product_id.name),'{}'.format(fournisseur)]
				]
				for forecast_id in production_schedule_id_forecast_ids:
					line[0] += [forecast_id['date_stop'].strftime("%d-%m-%Y"),forecast_id['date_stop'].strftime("%d-%m-%Y"),forecast_id['date_stop'].strftime("%d-%m-%Y"),forecast_id['date_stop'].strftime("%d-%m-%Y"),forecast_id['date_stop'].strftime("%d-%m-%Y")]
					line[1] += ['En stock','Demande planifiée','Réappro suggéré','Stock planifié','Entrepot ID']
					line[2].append(forecast_id['starting_inventory_qty'])
					line[2].append(forecast_id['forecast_qty'])
					line[2].append(forecast_id['replenish_qty'] )
					line[2].append(forecast_id['safety_stock_qty'])
					line[2].append(forecast_id.get('warehouse_id',[''])[0])
				header = True
			else:
				fournisseur = self.rechercher_fournisseur(production_schedule_id.product_id)
				line = [
					['{}'.format(production_schedule_id.product_id.default_code),'{}-{}'.format(production_schedule_id.product_id.default_code,production_schedule_id.product_id.name),'{}'.format(fournisseur)]
				]
				for forecast_id in production_schedule_id_forecast_ids:
					line[0].append(forecast_id['starting_inventory_qty'])
					line[0].append(forecast_id['forecast_qty'])
					line[0].append(forecast_id['replenish_qty'] )
					line[0].append(forecast_id['safety_stock_qty'])
					line[0].append(forecast_id.get('warehouse_id',[''])[0])
					
			export_xls.append(line)
			
			"""
			
			line = [[ '', ],['{}'.format(production_schedule_id.product_id.id),'{}-{}'.format(production_schedule_id.product_id.id,production_schedule_id.product_id.name),],['Demande planifiée'],['Réappro suggéré'],['Stock planifié'],['Entrepot ID']]
			for forecast_id in production_schedule_id_forecast_ids:
				line[0].append(forecast_id['date_stop'].strftime("%d-%m-%Y"))
				line[1].append(forecast_id['starting_inventory_qty'])
				line[2].append(forecast_id['forecast_qty'])
				line[3].append(forecast_id['replenish_qty'] )
				line[4].append(forecast_id['safety_stock_qty'])
				line[5].append(forecast_id.get('warehouse_id',[''])[0])
				
			export_xls.append(line)
			"""
		logger.info(export_xls)
		
		wb = Workbook()
		dest_filename = "/tmp/pdp.xlsx"
		ws1 = wb.active
		ws1.title = "PDP"		
		row = 1
		col = 1
		
		for line in export_xls:
			for l in line:
				for el in l:
					ws1.cell(row=row,column=col).value = el
					col +=1
				col = 1
				row += 1
			col = 1
			
		
		for column_cells in ws1.columns:
			
			length = max(len(cell.value and str(cell.value) or "") for cell in column_cells)
			ws1.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
		
		wb.save(filename = dest_filename)
		
		wb.close()
		
		with open(dest_filename, "rb") as xlfile:
			byte_data = xlfile.read()
		
		# on crée la pièce jointe
		attachment_obj = self.env['ir.attachment']
		
		# s'il en existe déjà une, on la supprime
		attachment = attachment_obj.search([('name','=','pdp.xlsx'),('res_model','=','pdp.export.wizard'),('res_id','=',self.ids[0])])
		attachment.unlink()
		
		value = {
			'type': 'binary',
			'datas': base64.b64encode(byte_data),
			'name': 'pdp.xlsx',
			'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
			'res_model': 'pdp.export.wizard',
			'res_id': self.ids[0],
		}
		
		attachment = attachment_obj.create(value)
		
		return {
			'type' : 'ir.actions.act_url',
			'url': '/web/content/{}/pdp.xlsx'.format(attachment.id),
			'target': 'self',
		}
				

	def rechercher_fournisseur(self,product):	
		nom_fournisseur = ' '
		if not product.seller_ids:
			nom_fournisseur = ' '
		else:
			for seller in product.seller_ids:
				code_fournisseur = seller.name
				nom_fournisseur = code_fournisseur.name
				break
		return nom_fournisseur			