# -*- coding: utf-8 -*-

from odoo import models, fields, api
import logging
logger = logging.getLogger(__name__)
import openpyxl
import io
import base64
from datetime import datetime

class wizard_pdp_import(models.TransientModel):
	_name = 'pdp.import.wizard'
	
	fichier = fields.Binary('Fichier à importer')
	
	
	def importer(self):
		
		# vérifier les données et le format de fichier
		data = base64.b64decode(self.fichier)
		output = open('/tmp/pdp_import.xlsx','wb')
		output.write(data)
		output.close()
				
		wb = openpyxl.load_workbook(filename='/tmp/pdp_import.xlsx')
		
		if 'PDP' not in wb.sheetnames:
			raise UserError('Aucun onglet du nom de PDP dans le fichier, import impossible')
		
		ws = wb['PDP']
		data = []
		for row in ws.iter_rows(min_row=1):
			line = []
			for cell in row:
				if type(cell.value) is str:
					line.append(cell.value.strip())
				else:
					line.append(cell.value)
			data.append(line)
		
		logger.info(data)
		lines = []
		for d in data:
			# on retire les lignes vides
			vide = True
			for c in d:
				if c:
					vide = False
			if not vide:
				lines.append(d)
		
		# 1ère ligne = dates
		# 2ème ligne = produit + qté stock
		# 3ème ligne = demande planifiée
		# 4ème ligne = réappro
		# 5ème ligne = stock planifié
		# 6ème ligne = warehouse_id
		
		i = 1
		dates = []
		demandes = []
		reappros = []
		
		# on vide ce qui est en cours
		self.env['mrp.production.schedule'].search([]).unlink()
		
		for line in lines:
			if i == 1:
				# on convertit les dates
				for j in range(1,len(line)):
					dates.append(datetime.strptime(line[j],'%d-%m-%Y'))
			if i == 2:
				# première colonne = nom produit
				product_id = self.env['product.product'].search([('id','=',line[0].split('-')[0])],limit=1)
				if len(product_id) == 0:
					raise UserError("Le produit dont le nom est {} n'a pas été trouvé dans Odoo".format(line[0]))
			if i == 3:
				# demandes planifiées
				for j in range(1,len(line)):
					demandes.append(line[j])
			if i == 4:
				# réappros
				for j in range(1,len(line)):
					reappros.append(line[j])
			if i == 6:
				# warehouse_id
				warehouse = self.env['stock.warehouse'].search([('id','=',line[1])])
				i = 0
				forecast_ids = []
				for j in range(len(dates)):
					value_forecast = {
						'date': dates[j],
						'forecast_qty': demandes[j],
						'replenish_qty': reappros[j],
					}
					forecast_ids.append((0,0,value_forecast))				
				
				value_schedule = {
					'product_id': product_id.id,
					'warehouse_id': len(warehouse) == 1 and warehouse.id or self.env['mrp.production.schedule']._default_warehouse_id().id,
					'forecast_ids': forecast_ids,
				}
				self.env['mrp.production.schedule'].create(value_schedule)
				
			i += 1
		
		
		return {
			'type': 'ir.actions.client',
			'tag': 'reload',
		}
